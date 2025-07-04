import os
from openpyxl import Workbook, load_workbook

def append_to_excel(file_path, headers, row):
    """
    Append a row to an Excel file.
    If the file doesn't exist, create it and write headers.
    """
    file_exists = os.path.exists(file_path)

    if file_exists:
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)

    sheet.append(row)
    workbook.save(file_path)
